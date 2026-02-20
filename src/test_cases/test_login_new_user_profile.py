import os
import time

from selenium.common import NoSuchElementException, TimeoutException, ElementNotInteractableException, StaleElementReferenceException
from selenium.webdriver.support.ui import Select


import pytest
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC





from pageObjects.user_details import UserDetails
from src.pageObjects.login_page import LoginPage
from src.utilities.logger import LogGen
from src.utilities.read_config import ReadConfig



# import pytest
#
from src.pageObjects.home_page import Home_page
from src.pageObjects.login_page import LoginPage
from src.utilities.logger import LogGen
from src.utilities.read_config import ReadConfig

import pandas as pd


from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv
from src.pageObjects.home_page import Home_page as Hp


def read_questions_from_excel(file_path):
    df = pd.read_excel(file_path)
    return df['Questions'].dropna().tolist()


def append_to_excel(file_path, question, answer=None, start_time=None, end_time = None, response_time =None):
    row_data = [question, answer, start_time, end_time, response_time]
    columns = ["Question", "Answer", "Start Time", "End Time", "Response Time (sec)"]

    if not os.path.exists(file_path):
        df = pd.DataFrame([row_data], columns=columns)
        df.to_excel(file_path, index=False)
    else:
        book = load_workbook(file_path)
        sheet = book.active
        sheet.append(row_data)
        for col_num, col_width in enumerate([60, 200, 25, 25, 20], start=1):
            sheet.column_dimensions[get_column_letter(col_num)].width = col_width
        book.save(file_path)
        book.close()


@pytest.mark.usefixtures("setup")
class TestVerificationCode:
    """Test verification code with negative, blank, and correct inputs"""

    baseURL = ReadConfig.getApplicationURL()
    email = ReadConfig.getEmail()
    verification_code = ReadConfig.get_verification_code()
    logger = LogGen.loggen()


    # def test_verification_code_all_scenarios(self):
    #     """Test blank, wrong, and correct verification codes"""
    #
    #     self.logger.info("\n" + "=" * 70)
    #     self.logger.info("VERIFICATION CODE TEST - ALL SCENARIOS")
    #     self.logger.info("=" * 70)
    #
    #     # ===== SCENARIO 1: BLANK CODE =====
    #     self.logger.info("\n[TEST 1] BLANK verification code...")
    #     self.driver.get(self.baseURL)
    #     time.sleep(2)
    #
    #     lp = LoginPage(self.driver)
    #     lp.click_login_with_verification_code()
    #     lp.set_email(self.email)
    #     lp.click_send_code()
    #     time.sleep(2)
    #
    #     lp.set_verification_code("")  # Blank
    #     time.sleep(1)
    #
    #     try:
    #         lp.click_login()
    #         self.logger.info("⚠ Checking if blank code is rejected...")
    #     except:
    #         self.logger.info("✓ Blank code - login button disabled")
    #
    #     time.sleep(2)
    #
    #     # ===== SCENARIO 2: WRONG CODE =====
    #     self.logger.info("\n[TEST 2] WRONG verification code (123456)...")
    #     self.driver.get(self.baseURL)
    #     time.sleep(2)
    #
    #     lp = LoginPage(self.driver)
    #     lp.click_login_with_verification_code()
    #     lp.set_email(self.email)
    #     lp.click_send_code()
    #     time.sleep(2)
    #
    #     lp.set_verification_code("123456")  # Wrong code
    #     time.sleep(1)
    #     lp.click_login()
    #     time.sleep(3)
    #
    #     if 'auth' in self.driver.current_url.lower():
    #         self.logger.info("✓ Wrong code rejected - still on login page")
    #     else:
    #         self.logger.warning("⚠ Wrong code was accepted (unexpected)")
    #
    #     # ===== SCENARIO 3: SPECIAL CHARACTERS =====
    #     self.logger.info("\n[TEST 3] INVALID code with special chars (!@#$%^)...")
    #     self.driver.get(self.baseURL)
    #     time.sleep(2)
    #
    #     lp = LoginPage(self.driver)
    #     lp.click_login_with_verification_code()
    #     lp.set_email(self.email)
    #     lp.click_send_code()
    #     time.sleep(2)
    #
    #     lp.set_verification_code("!@#$%^")  # Special characters
    #     time.sleep(1)
    #     lp.click_login()
    #     time.sleep(3)
    #
    #     if 'auth' in self.driver.current_url.lower():
    #         self.logger.info("✓ Special characters rejected")
    #
    #     # ===== SCENARIO 4: CORRECT CODE =====
    #     self.logger.info("\n[TEST 4] CORRECT verification code...")
    #     self.driver.get(self.baseURL)
    #     time.sleep(2)
    #
    #     lp = LoginPage(self.driver)
    #     lp.click_login_with_verification_code()
    #     lp.set_email(self.email)
    #     lp.click_send_code()
    #     time.sleep(2)
    #
    #     lp.set_verification_code(self.verification_code)  # CORRECT
    #     self.logger.info(f"Using correct code: {self.verification_code}")
    #     time.sleep(1)
    #     lp.click_login()
    #     time.sleep(5)
    #     print("hi")
    #
    #     #Verify success
    #     if 'auth' not in self.driver.current_url.lower():
    #         self.logger.info("✓ CORRECT code accepted - Login successful!")
    #         self.logger.info("\n" + "=" * 70)
    #         self.logger.info("✅ ALL SCENARIOS PASSED")
    #         self.logger.info("=" * 70)
    #         assert True
    #     else:c
    #         self.logger.error("❌ Correct code failed")
    #         self.driver.save_screenshot("./screenshots/correct_code_failed.png")
    #         assert False, "Login should succeed with correct code"
    #     print("hii")
    #     time.sleep(10)

    @pytest.mark.testcases
    def test_positive_login(self):
        self.driver.get(self.baseURL)
        time.sleep(10)
        self.driver.find_element(By.XPATH, "//button[@class='text-brand-blue py-3 px-4 mb-4 font-semibold text-sm']").click()
        time.sleep(5)
        print("hiii")
        self.driver.find_element(By.XPATH, "//input[@name='user_email']").send_keys("front.end@metadevelopers.in")
        time.sleep(3)
        print("hi1")
        self.driver.find_element(By.XPATH, "//button[@type='submit']").click()
        time.sleep(5)
        print("hi2")
        self.driver.find_element(By.XPATH, "//input[@type='text'][1]").send_keys("112233")
        time.sleep(5)
        print("hi3")
        self.driver.find_element(By.XPATH, "//button[@type='submit']").click()
        time.sleep(5)
        print("hi4")

        # self.driver.find_element(By.XPATH, "(//span[text()='User Management'])[2]").click()
        # self.driver.find_element(By.CSS_SELECTOR, "span.mr-2").click()
        # time.sleep(5)
        # self.Hp = Home_page(self.driver)
        # time.sleep(5)
        # self.get_mail = ReadConfig()
        # # self.driver.find_element(By.CSS_SELECTOR, "input.bg-Gray_G20").send_keys("abc@gmail.com")
        # # self.Hp.click_user_management()
        # # self.Hp.click_create_new_user()
        # acv = self.get_mail.get_new_email()
        # self.driver.find_element(By.CSS_SELECTOR, "input.bg-Gray_G20").send_keys(acv)
        # self.Hp.click_radio_user()
        # self.Hp.click_create_user()






# @pytest.mark.usefixtures("setup")
# class TestAufgang:
#     baseURL = ReadConfig.getApplicationURL()
#     verification_code = ReadConfig.get_verification_code()
#     new_email = ReadConfig.get_new_email()
#     logger = LogGen.loggen()
#     new_user_email = ReadConfig.get_new_email()
#     first_name = ReadConfig.get_first_name()
#     last_name = ReadConfig.get_last_name()
#     phone_number = ReadConfig.get_phone_number()
#     location = ReadConfig.get_location()
#     project_name = ReadConfig.get_project_name()
#     client_name = ReadConfig.get_client_name   ()
#     total_area = ReadConfig.get_total_area()
#     estimated_budget = ReadConfig.get_estimated_budget()
#     district_code = ReadConfig.get_district_code()
#     location_p = ReadConfig.get_location_p()
#
    # baseURL = ReadConfig.getApplicationURL()
    # email = ReadConfig.getEmail()
    # verification_code = ReadConfig.get_verification_code()
    # new_email = ReadConfig.get_new_email()
    # logger = LogGen.loggen()

    def test_Admin_Login(self):
        self.logger.info("Test_002_Create_User")
        self.logger.info("Starting_Test_create_new_user")

        #self.driver = setup
        # print("hi")
        # self.driver.get(self.baseURL)
        # print("hi1")
        # self.lp = LoginPage(self.driver)
        # self.lp.click_login_with_verification_code()
        # print("hiii")
        # self.lp.set_email(self.email)
        # self.lp.click_send_code()
        # self.lp.set_verification_code(self.verification_code)
        #
        # self.lp.click_login()
        #

        time.sleep(8)
        self.driver.find_element(By.XPATH, "(//span[text()='User Management'])[2]").click()
        self.driver.find_element(By.CSS_SELECTOR, "span.mr-2").click()
        time.sleep(5)
        self.Hp = Home_page(self.driver)
        time.sleep(5)
        self.get_mail = ReadConfig()
        # self.driver.find_element(By.CSS_SELECTOR, "input.bg-Gray_G20").send_keys("abc@gmail.com")
        # self.Hp.click_user_management()
        # self.Hp.click_create_new_user()
        acv = self.get_mail.get_new_email()
        self.driver.find_element(By.CSS_SELECTOR, "input.bg-Gray_G20").send_keys(acv)
        self.Hp.click_radio_user()
        self.Hp.click_create_user()
        # self.Hp = Home_page(self.driver)
        #
        # time.sleep(5)
        # self.driver.find_element(By.XPATH, "(//span[text()='User Management'])[2]").click()
        # self.driver.find_element(By.CSS_SELECTOR, "span.mr-2").click()
        # time.sleep(5)
        # self.get_mail = ReadConfig()
        # #self.driver.find_element(By.CSS_SELECTOR, "input.bg-Gray_G20").send_keys("abc@gmail.com")
        # # self.Hp.click_user_management()
        # #self.Hp.click_create_new_user()
        # self.get_mail.get_new_email()
        # self.Hp.click_radio_user()
        # self.Hp.click_create_user()

        print("h000")

        time.sleep(5)
        a = self.driver.find_elements(By.XPATH,"//span[@class='sm:hidden lg:flex font-semibold md:text-xs lg:text-sm']")
        for i in a:
            if i.text == "Projects":
                i.click()
                break
        time.sleep(5)
        self.driver.find_element(By.XPATH,"//button[@class='flex items-center bg-brand-blue text-white border rounded-md py-2 px-6 hover:bg-btn_hov_brand font-bold']").click()
        time.sleep(3)
        print("H1122")

        self.driver.find_element(By.XPATH,"// input[ @ name = 'project_name']").send_keys("Saaeendarssstone")
        self.driver.find_element(By.XPATH, "// input[ @ name = 'project_client_name']").send_keys("Nedd stark")
        self.driver.find_element(By.XPATH,"//input[@class='w-full relative ps-[80px] bg-Gray_G20 rounded-lg px-4 py-3 border border-grey-200 focus:outline-none mt-2 text-xs lg:text-sm refresh']").send_keys("2200")
        self.driver.find_element(By.XPATH,"//input[@class='w-full relative ps-[60px] rounded-lg px-4 py-3 border border-grey-200 focus:outline-none mt-2 text-xs lg:text-sm refresh']").send_keys("129278")
        self.driver.find_element(By.XPATH, "// input[ @ name = 'project_district_code']").send_keys("NHI")
        abcd = self.driver.find_element(By.XPATH, "// input[ @class ='w-full rounded-lg px-4 py-3 border border-grey-200 focus:outline-none mt-2 text-xs lg:text-sm refresh pac-target-input']")
        abcd.send_keys("New")
        time.sleep(3)
        abcd.send_keys(Keys.ARROW_DOWN)

        time.sleep(5)
        self.driver.find_element(By.XPATH,"//input[@placeholder='Enter base level']").send_keys("300")
        print("Hello")

        self.driver.find_element(By.XPATH,'(//input[@class="w-full relative ps-[80px] bg-Gray_G20 rounded-lg px-4 py-3 border border-grey-200 focus:outline-none mt-2 text-xs lg:text-sm refresh"])[2]').send_keys("220000")
        print("Hello11")
        time.sleep(10)
        self.driver.find_element(By.XPATH,"//button[@type='submit']").click()
        time.sleep(10)
        print("Hello2")
        self.driver.find_element(By.XPATH,"//input[@name='project_building_height']").send_keys("8000")
        self.driver.find_element(By.XPATH,"//input[@name='project_number_of_stories']").send_keys("15")
        time.sleep(20)
        # self.driver.find_element(By.XPATH,"//div[@class='flex justify-between items-center w-full bg-Gray_G20 rounded-lg px-4 py-[15px] border border-grey-200 focus:outline-none mt-2 text-xs lg:text-sm refresh cursor-pointer']").click()
        # time.sleep(10)
        # self.driver.find_element(By.XPATH,"//li[@class='px-4 w-[100%] py-2 text-xs cursor-pointer text-black-900 hover:bg-gray-100'][1]").click()
        # time.sleep(3)
        # self.driver.find_element(By.XPATH, "//*[@id='root']/div[1]/div[3]/main/div/div[3]/div/div[2]/div/div/form/div[1]/div/div/div[1]/div[2]/div/div/div/span").click()
        # # self.driver.find_element(By.XPATH, "//div[@class='flex justify-between items-center w-full bg-Gray_G20 rounded-lg px-4 py-[15px] border border-grey-200 focus:outline-none mt-2 text-xs lg:text-sm refresh cursor-not-allowed opacity-60']").click()
        # time.sleep(3)
        # self.driver.find_element(By.XPATH,"// li[ @class ='px-4 w-[100%] py-2 text-xs cursor-pointer text-black-900 hover:bg-gray-100'][1]").click()
        # time.sleep(3)
        # self.driver.find_element(By.XPATH,"//input[@name='project_building_height']").send_keys("8000")
        # time.sleep(3)
        # self.driver.find_element(By.XPATH,"// input[ @ name = 'project_number_of_stories']").send_keys("12")
        # time.sleep(3)

        calender_icons = self.driver.find_elements(By.XPATH, "//img[contains(@alt,'calendar')]")
        dates_to_select = [{"year": "2025", "month": "December", "day": "15"},
                           {"year": "2026", "month": "December", "day": "16"}]

        for i in range(len(calender_icons)):
            calender_icons[i].click()
            year_to_select = dates_to_select[i]["year"]
            month_to_select = dates_to_select[i]["month"]
            day_to_select = dates_to_select[i]["day"]

            year_picker = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//span[@class='rdrYearPicker']")))
            year_picker.click()

            year_option = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, f"//option[text()='{year_to_select}']")))
            year_option.click()

            month_picker = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//span[contains(@class,'rdrMonthPicker')]")))
            month_picker.click()

            month_option = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, f"//option[text()='{month_to_select}']")))
            month_option.click()

            day_button = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(
                (By.XPATH, f"//span[contains(@class,'rdrDayNumber')]//span[text()='{day_to_select}']")))
            day_button.click()
            time.sleep(2)
            print("Hello3")

        #########################################################
        # self.driver.find_element(By.XPATH,"//div[@class='w-full relative'][1]").click()
        # time.sleep(10)
        # ababba = self.driver.find_elements(By.XPATH,'//div[@class="flex items-center rounded-lg px-3 w-full py-2 text-sm cursor-pointer text-gray-900 hover:bg-gray-100 transition-colors"]')
        # ababba[1].click()
        # time.sleep(10)
        # ababba[2].click()
        # time.sleep(10)
        # self.driver.find_element(By.XPATH, "//div[@class='w-full relative'][1]").click()
        # time.sleep(5)
        # self.driver.find_element(By.XPATH, "//div[@class='flex flex-wrap gap-1 justify-start items-center w-full px-2 py-2 min-h-[46px] bg-white border border-grey-200 rounded-md focus:outline-none text-xs lg:text-sm cursor-pointer transition-colors'][1]").click()
        # time.sleep(20)
        # self.driver.find_element(By.XPATH, "//div[@class='flex justify-between items-center w-full bg-Gray_G20 rounded-lg px-4 py-[15px] border border-grey-200 focus:outline-none mt-2 text-xs lg:text-sm refresh cursor-pointer']").click()
        # time.sleep(20)
        # self.driver.find_element(By.XPATH, "//li[@class='px-4 w-[100%] py-2 text-xs cursor-pointer text-black-900 hover:bg-gray-100'][4]").click()
        # time.sleep(10)
        # self.driver.find_element(By.XPATH, "//div[@class='flex flex-wrap gap-1 justify-start items-center w-full  rounded-lg px-2 py-2 border border-gray-200 focus:outline-none mt-2 text-xs lg:text-sm cursor-pointer']").click()
        # time.sleep(10)
        # self.driver.find_element(By.XPATH, "//div[@class='text-gray-700 select-none cursor-pointer']").click()
        # time.sleep(10)
        ###########################################################333

        a = self.driver.find_elements(By.XPATH, "//div[@class='w-full relative'][1]")
        for i in range(len(a)):
            if i == 0:
                a[i].click()
                ababba = self.driver.find_elements(By.XPATH,'//div[@class="flex items-center rounded-lg px-3 w-full py-2 text-sm cursor-pointer text-gray-900 hover:bg-gray-100 transition-colors"]')
                ababba[1].click()
                time.sleep(5)
                ababba[2].click()
                time.sleep(5)
                a[i].click()
            if i ==1:
                a[i].click()
                time.sleep(5)
                xttt= self.driver.find_elements(By.XPATH,"//div[@class='flex items-center rounded-lg px-3 w-full py-2 text-sm cursor-pointer text-gray-900 hover:bg-gray-100 transition-colors']")
                xttt[0].click()
                xttt[1].click()
                a[i].click()

                time.sleep(20)
                # self.driver.find_elements(By.XPATH,"//span[@class='text-xs text-black']").click()
                # time.sleep(5)
                # self.driver.find_element(By.XPATH, "//div[@class='flex justify-between items-center w-full bg-Gray_G20 rounded-lg px-4 py-[15px] border border-grey-200 focus:outline-none mt-2 text-xs lg:text-sm refresh cursor-pointer']").click()
                # time.sleep(20)
                self.driver.find_element(By.XPATH, "//img[@class='transition-transform duration-300 me-2 absolute right-2 lg:right-0 top-1/2 transform -translate-y-1/2 cursor-pointer  lg:w-3 lg:h-3 w-2 h-2']").click()
                time.sleep(20)
                self.driver.find_element(By.XPATH, "//li[@class='px-4 w-[100%] py-2 text-xs cursor-pointer text-black-900 hover:bg-gray-100'][4]").click()
                time.sleep(10)
                self.driver.find_element(By.XPATH, "//div[@class='flex flex-wrap gap-1 justify-start items-center w-full min-h-[46px] rounded-lg px-2 py-2 border border-gray-200 focus:outline-none mt-2 text-xs lg:text-sm cursor-pointer']").click()
                time.sleep(10)
                bbcdad = self.driver.find_elements(By.XPATH, '//input[@type="checkbox"]')
                bbcdad[1].click()
                time.sleep(5)
                bbcdad[2].click()
                time.sleep(5)






        # calender_icons = self.driver.find_elements(By.XPATH, "//img[contains(@alt,'calendar')]")
        #
        # for i in range(len(calender_icons)):
        #     # Click on the calendar icon (start date or end date)
        #     calender_icons[i].click()
        #
        #     # Wait for the year picker to be visible and clickable
        #     year_picker = WebDriverWait(self.driver, 10).until(
        #         EC.element_to_be_clickable((By.XPATH, "//select[@class='rdrYearPicker']"))
        #     )
        #
        #     # Use Select to interact with the year dropdown
        #     select_year = Select(year_picker)
        #     select_year.select_by_visible_text('2025')  # Change to desired year
        #
        #     # Wait for the month dropdown to be visible and clickable
        #     month_picker = WebDriverWait(self.driver, 10).until(
        #         EC.element_to_be_clickable((By.XPATH, "//span[contains(@class,'rdrMonthPicker')]"))
        #     )
        #     month_picker.click()
        #
        #     # Wait for and select December from the month options (you can change the month as needed)
        #     month_option = WebDriverWait(self.driver, 10).until(
        #         EC.element_to_be_clickable((By.XPATH, "//span[text()='December']"))
        #     )
        #     month_option.click()
        #
        #     # Wait for the day button (for the 1st day in the calendar) to be clickable
        #     day_button = WebDriverWait(self.driver, 10).until(
        #         EC.element_to_be_clickable((By.XPATH, "//span[contains(@class,'rdrDayNumber')]//span[text()='1']"))
        #     )
        #     day_button.click()
        #
        #     # Optional: Wait before moving on to the next calendar if necessary
        #     time.sleep(2)
        #
        #     print(f"Clicked on calendar icon {i + 1} and selected the date.")
        #
        #     # Allow some time for UI updates
        #     time.sleep(2)










            # month = self.driver.find_element(By.XPATH, "//span[@class='rdrMonthPicker']")
            # month.click()
            #
            # month_element = self.driver.find_element(By.XPATH, "//td[@class='rdtMonth' and text()='Jun']")
            # month_element.click()
            # day_element = self.driver.find_element(By.XPATH, "//td[@class='rdtDay' and text()='21']")
            # day_element.click()





        # self.logger.info("Selecting start date...")
        #
        # # Click calendar icon (not the input field!)
        # calendar_icons = self.driver.find_elements(By.XPATH, "//img[@alt='calendar']")
        # calendar_icons[0].click()  # First icon = start date
        # time.sleep(2)
        #
        # Select(self.driver.find_element(By.XPATH, "//select[contains(@class,'rdrYearPicker')]")).select_by_visible_text(
        #     '2025')
        # time.sleep(1)
        #
        # Select(self.driver.find_element(By.XPATH, "//select[@class='rdrMonthPicker']")).select_by_visible_text(
        #     'December')
        # time.sleep(1)
        #
        # self.driver.find_element(By.XPATH, "//button[@class='rdrDay']//span[text()='1']").click()
        # time.sleep(2)
        #
        # # COMPLETION DATE
        # self.logger.info("Selecting completion date...")
        #
        # # Click second calendar icon
        # calendar_icons = self.driver.find_elements(By.XPATH, "//img[@alt='calendar']")
        # calendar_icons[1].click()  # Second icon = completion date
        # time.sleep(2)
        #
        # Select(self.driver.find_element(By.XPATH, "//select[contains(@class,'rdrYearPicker')]")).select_by_visible_text(
        #     '2026')
        # time.sleep(1)
        #
        # Select(self.driver.find_element(By.XPATH, "//select[@class='rdrMonthPicker']")).select_by_visible_text('March')
        # time.sleep(1)
        #
        # self.driver.find_element(By.XPATH, "//button[@class='rdrDay']//span[text()='31']").click()
        # time.sleep(2)
        #
        # Proceed

        self.driver.find_element(By.XPATH, "//button[text()='Proceed']").click()
        time.sleep(10)


        self.driver.find_element(By.XPATH,"//input[@placeholder='Enter Revit GUID (must be 36 characters)']").send_keys("e398c797-9a5c-4b45-a1d9-5534223bd9a8")
        self.driver.find_element(By.XPATH, "//input[@placeholder='Enter the link']").send_keys("https://zola.planning.nyc.gov/l/lot/1/1928/48#18.24/40.808055/-73.950041")
        self.driver.find_element(By.XPATH, "//span[@class='text-xs lg:text-sm']").click()
        bbb =self.driver.find_element(By.XPATH, "//span[@class='text-gray-400 text-xs px-2 py-1 ps-3']")
        bbb.click()
        dadccb = self.driver.find_elements(By.XPATH, '//input[@type="checkbox"]')
        dadccb[1].click()
        time.sleep(5)
        dadccb[2].click()

        self.driver.find_element(By.XPATH, "//div[@class='flex flex-wrap gap-1 rounded-lg px-1 w-full py-2 border border-grey-200 focus:outline-none cursor-pointer']").click()
        print("hjiiiii")
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//textarea[@name='project_description']").send_keys("Describe all")
        self.driver.find_element(By.XPATH, "//button[@class='text-cool-gray flex space-x-1 items-center font-semibold px-4 py-2 rounded-md border border-brand-blue text-xs lg:text-sm cursor-pointer']").click()
        time.sleep(20)
        self.driver.find_element(By.XPATH, "//button[@type='submit']").click()
        time.sleep(10)
        print("last")
        time.sleep(5)
        # self.driver.find_element(By.XPATH, "(//span[@class='ml-2'])[2]").click()
        # self.driver.find_element(By.XPATH, "//button[@class='flex justify-center pt-4 cursor-pointer text-[16px] mt-2 font-semibold hover:text-hover-brand text-center']").click()
        #self.driver.quit()
        print("lasstt")



    @pytest.mark.testcases
    def test_Elise_AI(self):
        try:
            self.driver.find_element(By.XPATH, "//button[contains(text(),'Elise AI')]").click()
        except NoSuchElementException:
            raise Exception("'Elise AI' button not found")
        try:
            WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//input[@name='chat-input']"))
            )
        except TimeoutException:
            raise Exception("Chatbot textarea not found")

        input_file = "C:\\AufGang_Python\\Questions.xlsx"
        output_file = "C:\\AufGang_Python\\Answers.xlsx"

        questions = read_questions_from_excel(input_file)

        for question in questions:
            response_text = ""

            for attempt in range(3):
                try:
                    textarea = WebDriverWait(self.driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, "//input[@name='chat-input']"))
                    )
                    self.driver.execute_script("arguments[0].scrollIntoView(true);", textarea)
                    textarea.clear()
                    textarea.send_keys(question)
                    break
                except (ElementNotInteractableException, StaleElementReferenceException, TimeoutException):
                    time.sleep(1)

            try:
                self.driver.find_element(By.XPATH, "//button[contains(@class, 'sendButton')]").click()
                time.sleep(15)
            except NoSuchElementException:
                response_text = "Send button not found"
                continue

            start_dt = datetime.now()
            start_time_str = start_dt.strftime("%Y-%m-%d %H:%M:%S")
            start_epoch = time.time()

            WebDriverWait(self.driver, 600).until(
                lambda d: (len(d.find_elements(By.XPATH, "(//div[contains(@class,'_messageBubbleAI1_1pctf_369')])[last()]//span")) == 0))

            response_duration = round(time.time() - start_epoch, 2)
            end_dt = datetime.now()
            end_time_str = end_dt.strftime("%Y-%m-%d %H:%M:%S")

            try:
                for attempt in range(3):
                    try:
                        time.sleep(2)

                        combined_content = []

                        p_elements = self.driver.find_elements(By.XPATH,"(//div[contains(@class, '_messageBubbleAI1_1pctf_369')])[last()]//p")
                        for p in p_elements:
                            text = p.text.strip()
                            if text:
                                combined_content.append(text)

                        time.sleep(5)

                        table_rows = self.driver.find_elements(By.XPATH,
                                                               "(//div[contains(@class, 'aiMessage') and contains(@class, 'messageWrapper')])[last()]//table//tr")
                        if table_rows:
                            for row in table_rows:
                                headers = row.find_elements(By.TAG_NAME, "th")
                                cells = row.find_elements(By.TAG_NAME, "td")

                                if headers:
                                    row_data = [h.text.strip() for h in headers]
                                else:
                                    row_data = [c.text.strip() for c in cells]

                                combined_content.append(" | ".join(row_data))

                        unique_lines = list(dict.fromkeys(combined_content))
                        response_text = "\n".join(unique_lines)

                        if response_text:
                            break
                        else:
                            print("No paragraph or table found")

                    except StaleElementReferenceException:
                        time.sleep(1)
                        if attempt == 2:
                            response_text = "Stale element - retry failed"

            except:
                response_text = "No response or element not found"

            append_to_excel(output_file, question, response_text, start_time_str, end_time_str, response_duration)
            time.sleep(10)































